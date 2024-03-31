import openpyxl
import re
import random

excel1 = openpyxl.load_workbook('Python2.xlsx')
sheet1 = excel1['Sheet1']
# excel2 = openpyxl.load_workbook('成绩2.xlsx')
with open('info.txt', 'r', encoding='utf8') as f:
    info: dict = eval(f.read())

# sheet2 = excel2['Sheet1']

index1 = 2
# index2 = 2

# name = 1
# no = 2
# py = 3
# ai = 4

# py_score_regex = re.compile(r'(.*?)分（Python结课考试）')
# ai_score_regex = re.compile(r'(.*?)分（人工智能通识课结课考试）')

# res = py_score_regex.match('100.0分（人工智能通识课结课考试）').groups()
# print(res)
# info = {}
for row in sheet1.iter_rows(min_row=2, values_only=True):
    # if not row:
    #     continue
    zuoye = random.randint(85, 90)
    shixun = 80
    sheet1.cell(index1, 6, zuoye)
    sheet1.cell(index1, 7, shixun)

    name = row[1]
    if name is not None:
        scores = info.get(name)
        if scores is None:
            index1 += 1
            continue
        score = scores.get('py')
        sheet1.cell(index1, 5, int(float(score)))
    index1 += 1

    # name = row[0]
    # no = row[1]
    # py = row[2]
    # ai = row[3]
    # info[name] = {'py': py if py else 0, 'ai': ai if ai else 0}

# print(info)
# index1 += 1
#     py_score = 0
#     ai_score = 0
#     if not row[0]:
#         break
#     _score = row[2]
#     if _score is not None:
#         scores = _score.split('\n')
#         sheet2.cell(index2, name, row[0])
#         sheet2.cell(index2, no, row[1])
#         for score in scores:
#             if res1 := py_score_regex.match(score):
#                 sheet2.cell(index2, py, res1.groups()[0])
#             if res2 := ai_score_regex.match(score):
#                 sheet2.cell(index2, ai, res2.groups()[0])
#         index2 += 1
# break

# a = random.randint(83, 90)
# b = random.randint(85, 87)
# c = int(a * 0.4 + b * 0.6)
#
# sheet.cell(index, 5, a)
# sheet.cell(index, 6, b)
#     sheet.cell(index, 7, s3)
#     sheet.cell(index, 8, res)
# index1 += 1
# break

excel1.save('Python3.xlsx')
# excel2.save('成绩2.xlsx')
excel1.close()
# excel2.close()

# for i in range(100):
#     a = random.randint(83, 90)
#     b = random.randint(85, 87)
#     c = int(a * 0.4 + b * 0.6)
#
#     print(a)
#     print(b)
#     print(c)
#     print()
# import re

# score = '73.0分（深度学习基础结课考试）\n76.0分（深度学习高级结课考试）\n75.0分（人工智能通识课结课考试）\n100.0分（Python结课考试）'
# regex = re.compile(r'(\d+\.?\d*)分')
# res = regex.findall(score)
# print(res)
# print(score.split('\n'))
